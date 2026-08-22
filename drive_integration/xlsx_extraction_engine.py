# -*- coding: utf-8 -*-
"""
xlsx_extraction_engine.py — Sales Tracking & Financial Spreadsheet Parser
"""

from pathlib import Path
from typing import Dict, Any, List

logger = __import__("logging").getLogger("nexa.extract.xlsx")


class XLSXExtractionEngine:
    def extract_sheets_and_tables(self, file_path: Path) -> Dict[str, Any]:
        """Extracts sheets, columns, and rows from XLSX / CSV files."""
        if not file_path.exists():
            return {"sheets": {}, "raw_text": "", "confidence": 0.0, "success": False}

        sheets_data = {}
        text_lines = []
        ext = file_path.suffix.lower()

        try:
            if ext in ('.xlsx', '.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    sheet_rows = []
                    text_lines.append(f"--- Tablo: {sname} ---")
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                        if cells:
                            sheet_rows.append(cells)
                            text_lines.append(" | ".join(cells))
                    sheets_data[sname] = sheet_rows
                wb.close()
            elif ext == '.csv':
                import csv
                with open(file_path, newline='', encoding='utf-8', errors='replace') as f:
                    reader = csv.reader(f)
                    csv_rows = [[c.strip() for c in r if c.strip()] for r in reader if r]
                    sheets_data["Default"] = csv_rows
                    for r in csv_rows:
                        text_lines.append(" | ".join(r))

            full_text = "\\n".join(text_lines)
            confidence = 98.0 if len(full_text) > 100 else 60.0
            return {
                "sheets": sheets_data,
                "raw_text": full_text,
                "confidence": confidence,
                "success": bool(full_text)
            }
        except Exception as e:
            logger.warning("Spreadsheet extraction error (%s): %s", file_path.name, e)
            return {"sheets": {}, "raw_text": "", "confidence": 0.0, "success": False, "error": str(e)}
