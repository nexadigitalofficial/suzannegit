"""
Otonom dosya izleyici: projeler/ klasöründeki PDF/video değişikliklerini
algılar, RAG dokümanlarını (NEXA PRIME DB) otomatik günceller ve importer'ı
tetikler. Kullanıcı müdahalesiz "dosya -> DB -> site -> ALYA" zinciri.
"""
import os
import re
import sys
import json
import time
import logging
import sqlite3
from pathlib import Path
import PyPDF2

logger = logging.getLogger("nexa.watchdog")

SITE_DIR = Path(__file__).resolve().parent
PROJELER_DIR = SITE_DIR / "projeler"
STATE_FILE = SITE_DIR / "watch_state.json"
DB = SITE_DIR / "nexa_database.db"
GARBAGE = re.compile(r"[\u0080-\u02FF]{4,}")

FOLDER_TO_DB = {
    '01_CONCEPT_BULVAR': 'CONCEPT BULVAR', '02_ANGIM_BEYTEPE': 'ANGİM BEYTEPE',
    '03_ANKAPORT_SARAY': 'ANKAPORT - SARAY', '04_GRANDE_YASAMKENT': 'GRANDE YAŞAMKENT',
    '05_GOKDEMIR_IMZA': 'GÖKDEMİR İMZA', '06_MONZA_EYLUL_CONCEPT': 'MONZA EYLÜL CONCEPT - VIP ÇAKIRLAR',
    '07_MONZA_MOON': 'MONZA MOON', '08_NARIN_RONYA_CITY': 'NARÇİN RONYA CITY - 1 (VIP WEST)',
    '09_TRIOLE_YASAM': 'TRIOLE YAŞAM', '10_EVART_YALIKAVAK': 'NEXA Royal Yalıkavak',
    '11_BORDO_YASAM': 'BORDO YAŞAM', '12_EXCELANCE_VADI': 'EXCELANCE VADİ',
    '13_EXCELANCE_BEYTEPE': 'EXCELANCE BEYTEPE', '14_GOKDEMIR_STAR': 'VIP YAŞAMKENT - GÖKDEMİR STAR',
    '15_IDEA_START_BRAVO': 'IDEA - START BRAVO', '16_JOVEN_PORT': 'JOVEN PORT',
    '17_JOVEN_KAMPUS': 'JOVEN KAMPÜS', '18_NEST_INCEK': 'NEST İNCEK',
    '19_NATURA_GOLF': 'NATURA GOLF', '20_NEVA_START_BRAVO': 'NEVA - START BRAVO',
    '21_SMD_TWIN': 'SMD TWIN', '22_SMD_PROTOKOL': 'SMD PROTOKOL',
    '23_S_POINT_SARAY': 'S POINT - VIP SARAY', 'ANGİM BEYTEPE': 'ANGİM BEYTEPE',
    'ANKAPORT - SARAY': 'ANKAPORT - SARAY', 'EVART YALIKAVAK': 'NEXA Royal Yalıkavak',
    'GRANDE YAŞAMKENT': 'GRANDE YAŞAMKENT', 'GÖKDEMİR İMZA': 'GÖKDEMİR İMZA',
    'IDEA - START BRAVO': 'IDEA - START BRAVO',
    'MONZA EYLÜL CONCEPT - VIP ÇAKIRLAR': 'MONZA EYLÜL CONCEPT - VIP ÇAKIRLAR',
    'MONZA MOON': 'MONZA MOON', 'NARÇİN RONYA CITY - 1 (VIP WEST)': 'NARÇİN RONYA CITY - 1 (VIP WEST)',
    'NEVA - START BRAVO': 'NEVA - START BRAVO', 'S POINT - VIP SARAY': 'S POINT - VIP SARAY',
    'TRIOLE YAŞAM': 'TRIOLE YAŞAM', 'VERDE MONA': 'VERDE MONA',
    'VIP AKADEMİ': 'VIP AKADEMİ', 'VIP AKADEMİ 2': 'VIP AKADEMİ 2',
    'VIP MARIN': 'VIP MARIN', 'VIP YAŞAMKENT - GÖKDEMİR STAR': 'VIP YAŞAMKENT - GÖKDEMİR STAR',
    'VIP YENİKENT': 'VIP YENİKENT', 'VIP ÇAKIRLAR': 'VIP ÇAKIRLAR',
    'VIP ÜNİVERSİTE': 'VIP ÜNİVERSİTE', 'VIVA - START BRAVO': 'VIVA - START BRAVO',
    'WM - PRIME': 'WM - PRIME',
}
NEW_PROJECTS = ['CONCEPT BULVAR', 'BORDO YAŞAM', 'EXCELANCE VADİ', 'EXCELANCE BEYTEPE',
                'JOVEN PORT', 'JOVEN KAMPÜS', 'NEST İNCEK', 'NATURA GOLF', 'SMD TWIN', 'SMD PROTOKOL']


def _extract_text(path):
    """Universal text extractor: PDF, XLSX, XLS, CSV, DOCX, TXT, MD."""
    ext = path.suffix.lower()
    text_content = []
    try:
        if ext == '.pdf':
            import contextlib
            import io
            with contextlib.redirect_stderr(io.StringIO()):
                r = PyPDF2.PdfReader(str(path))
                text_content = [(pg.extract_text() or "") for pg in r.pages]
        elif ext in ('.xlsx', '.xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    text_content.append(f"--- Sayfa / Tablo: {sheet} ---")
                    for row in ws.iter_rows(values_only=True):
                        row_data = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                        if row_data:
                            text_content.append(" | ".join(row_data))
                wb.close()
            except Exception as xe:
                logger.debug("xlsx extraction fallback: %s", xe)
        elif ext == '.csv':
            import csv
            with open(path, newline='', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f)
                for row in reader:
                    if row:
                        text_content.append(" | ".join(c.strip() for c in row if c.strip()))
        elif ext in ('.docx', '.doc'):
            try:
                import docx
                doc = docx.Document(str(path))
                text_content = [p.text for p in doc.paragraphs if p.text.strip()]
            except Exception:
                try:
                    import zipfile
                    import xml.etree.ElementTree as ET
                    with zipfile.ZipFile(str(path)) as z:
                        xml_content = z.read('word/document.xml')
                        tree = ET.fromstring(xml_content)
                        text_content = [node.text for node in tree.iter() if node.text and node.text.strip()]
                except Exception:
                    pass
        elif ext in ('.txt', '.md', '.json'):
            text_content.append(path.read_text(encoding='utf-8', errors='replace'))
    except Exception as e:
        logger.warning("Metin çıkarma hatası %s: %s", path.name, e)
        return ""
    return "\n".join(text_content)


def _cat_for(title):
    t = title.upper()
    if any(k in t for k in ('SÖZLEŞME', 'SOZLESME', 'ŞARTNAME', 'SARTNAME', 'PROTOKOL', 'HUKUK', 'TAPU')):
        return 'SÖZLEŞME'
    if any(k in t for k in ('SATIŞ TAKİP', 'SATIS TAKIP', 'SATIŞ', 'SATIS', 'LİSTE', 'STOK', 'BLOK')):
        return 'SATIŞ TAKİP'
    if any(k in t for k in ('FİYAT', 'FIYAT', 'ÖDEME', 'ODEME', 'TABLO', 'TAKSİT', 'PEŞİNAT')):
        return 'FİYAT TABLOSU'
    if any(k in t for k in ('SUNUM', 'KATALOG', 'BROŞÜR', 'BROSUR', 'LANSMAN')):
        return 'SUNUM'
    if any(k in t for k in ('KAT', 'PLAN', 'VAZİYET', 'VAZIYET', 'KROKİ', 'MİMARİ', 'MIMARI')):
        return 'KAT PLANI'
    return 'GENEL'


def _chunk_text(text, size=1800, overlap=150):
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) < 90 or GARBAGE.search(text):
        return []
    out, i = [], 0
    while i < len(text):
        out.append(text[i:i + size])
        i += size - overlap
    return out


def _snapshot():
    snap = {}
    if not PROJELER_DIR.exists():
        return snap
    for folder in PROJELER_DIR.iterdir():
        if not folder.is_dir():
            continue
        for f in folder.iterdir():
            if f.is_file():
                snap[str(f)] = (f.stat().st_size, int(f.stat().st_mtime))
    return snap


def _normalize_title(folder_name):
    """'25_YENI_ORNEK_PROJE' -> 'YENİ ÖRNEK PROJE'"""
    t = re.sub(r'^\d+[_\s-]*', '', folder_name).strip()
    t = t.replace('_', ' ')
    t = re.sub(r'\s+', ' ', t).strip()
    return t or folder_name


def _ensure_project(folder_name, cur, name_to_id, map_data, order_data):
    """Yeni klasor => DB projesi + site karti + sira kaydi. (Kendini duzenleyen site)"""
    title = _normalize_title(folder_name)
    if title in name_to_id:
        pid = name_to_id[title]
    else:
        cur.execute("INSERT INTO projects (name, is_portfolio, listing_type, created_at) VALUES (?,0,'Satılık',datetime('now'))", (title,))
        pid = cur.lastrowid
        name_to_id[title] = pid
        logger.info("OTO-PROJE: %s -> id %s", title, pid)
    if any(m.get('title') == title or m.get('folder_name') == folder_name for m in map_data):
        return pid
    folder = PROJELER_DIR / folder_name
    pdfs = sorted(folder.glob('*.pdf')) if folder.exists() else []
    mp4s = sorted(folder.glob('*.mp4')) if folder.exists() else []
    used = [int(m['id'].split('-')[-1]) for m in map_data if m.get('id', '').startswith('cbvip-prj-')]
    next_no = (max(used) + 1) if used else 1
    card_id = f'cbvip-prj-{next_no}'
    map_data.append({
        'id': card_id, 'db_id': pid, 'title': title, 'folder_name': folder_name,
        'folder_path': f'projeler/{folder_name}',
        'has_presentation': len(pdfs) > 0, 'has_video': len(mp4s) > 0,
        'media_mode': 'local',
        'presentations': [{'filename': x.name, 'url': f'/stream/pdf/{card_id}'} for x in pdfs[:8]],
        'videos': [{'filename': x.name, 'url': f'/stream/video/{card_id}'} for x in mp4s[:6]],
        'drive_pdf_preview': '', 'drive_video_preview': '',
        'price_display': '', 'price': '', 'location': '', 'room_info': '',
    })
    logger.info("OTO-KART: %s -> %s", title, card_id)
    top = max((o.get('rank', 0) for o in order_data), default=0) + 1
    order_data.append({'id': card_id, 'title': title, 'rank': top,
                       'is_pinned': False, 'is_hidden': False})
    return pid


def ingest_changed(only_ingest=False):
    """Değişen PDF'leri DB'ye işler; importer'ı tetikler."""
    snap = _snapshot()
    state = {}
    if STATE_FILE.exists():
        try:
            state = json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            state = {}
    changed = [k for k, v in snap.items() if state.get(k) != v]
    if not changed:
        return 0, 0, "degisiklik yok"

    db = sqlite3.connect(DB, timeout=30)
    cur = db.cursor()
    name_to_id = {r[0]: r[1] for r in cur.execute("SELECT name, id FROM projects").fetchall()}
    for n in NEW_PROJECTS:
        if n not in name_to_id:
            cur.execute("INSERT INTO projects (name, is_portfolio, listing_type, created_at) VALUES (?,0,'Satılık',datetime('now'))", (n,))
            name_to_id[n] = cur.lastrowid
    added_docs = added_chunks = 0
    map_data = json.loads((SITE_DIR / "projects_map.json").read_text(encoding="utf-8")) if (SITE_DIR / "projects_map.json").exists() else []
    order_data = json.loads((SITE_DIR / "display_order.json").read_text(encoding="utf-8")) if (SITE_DIR / "display_order.json").exists() else []
    new_cards = False
    for key in changed:
        path = Path(key)
        ext = path.suffix.lower()
        if ext not in ('.pdf', '.xlsx', '.xls', '.csv', '.docx', '.doc', '.txt', '.md', '.json'):
            continue
        folder_name = path.parent.name
        pid = name_to_id.get(FOLDER_TO_DB.get(folder_name, ''))
        if pid is None:
            pid = _ensure_project(folder_name, cur, name_to_id, map_data, order_data)
            if pid is None:
                continue
            new_cards = True
        if pid is None:
            continue
        text = _extract_text(path)
        if len(text) < 40:
            continue
        title = path.stem
        doc_type = ext.replace('.', '')
        category = _cat_for(title)
        cur.execute("SELECT id FROM documents WHERE project_id=? AND title=? AND doc_type=?", (pid, title, doc_type))
        row = cur.fetchone()
        if row:
            cur.execute("DELETE FROM document_chunks WHERE document_id=?", (row[0],))
            cur.execute("UPDATE documents SET content=?, file_url=?, category=? WHERE id=?",
                        (text, str(path), category, row[0]))
            did = row[0]
        else:
            cur.execute("INSERT INTO documents (project_id, doc_type, title, content, file_url, category, created_at) VALUES (?,?,?,?,?,?,datetime('now'))",
                        (pid, doc_type, title, text, str(path), category))
            did = cur.lastrowid
            added_docs += 1
        for c in _chunk_text(text):
            cur.execute("INSERT INTO document_chunks (document_id, chunk_text) VALUES (?,?)", (did, c))
            added_chunks += 1
    db.commit()
    db.close()
    if new_cards:
        (SITE_DIR / "projects_map.json").write_text(json.dumps(map_data, ensure_ascii=False, indent=1), encoding="utf-8")
        (SITE_DIR / "display_order.json").write_text(json.dumps(order_data, ensure_ascii=False, indent=1), encoding="utf-8")
    STATE_FILE.write_text(json.dumps(snap, ensure_ascii=False, indent=1), encoding="utf-8")
    if not only_ingest and (added_docs or added_chunks):
        _run_importer()
    return added_docs, added_chunks, f"{len(changed)} dosya incelendi"


def _run_importer():
    try:
        import subprocess
        kwargs = {}
        if sys.platform == "win32" and hasattr(subprocess, "CREATE_NO_WINDOW"):
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        subprocess.Popen([sys.executable, str(SITE_DIR / "nexa_data_importer.py")],
                         cwd=str(SITE_DIR), **kwargs)
    except Exception:
        pass


def watchdog_loop(interval=120):
    logger.info("Dosya izleyici basladi (projeler/ klasoru, %s sn)", interval)
    while True:
        try:
            ingest_changed()
        except Exception as e:
            logger.warning("watchdog: %s", e)
        time.sleep(interval)


if __name__ == "__main__":
    print(ingest_changed())